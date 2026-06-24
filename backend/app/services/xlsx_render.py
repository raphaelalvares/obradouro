"""Renderização do cronograma da obra em Excel (.xlsx) — espelha o gráfico de Gantt da tela.

Função PURA (sem DB/IO): recebe a árvore já carregada (get_tree) + identidade e devolve os bytes do
.xlsx. As barras são preenchimento ESTÁTICO calculado AQUI — não dependem de o Excel recalcular
fórmula/formatação condicional; ou seja, 100% alimentado pelo sistema. Layout inspirado no
"Planejador de projeto" da Microsoft: faixa de meses, "Semana N" (7 dias), inicial do dia da semana
e barras coloridas por situação (mesmas cores da tela: previsto=âmbar, concluído=verde, atrasado=
vermelho; a fração preenchida = avanço). A seleção de linhas espelha gantt.ts (etapa-resumo + suas
tarefas com datas) → o que se vê na tela é o que se baixa.
"""

import datetime as dt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# Cores em ARGB (FF + RGB). Iguais às da tela do Gantt — leem bem na tela e impressas.
GOLD = "FFD8A53A"
GOLD_DONE = "FFA9801F"
LATE = "FFE5654B"
LATE_DONE = "FFB14430"
DONE = "FF5FB87A"
WEEKEND = "FFF4F2EC"
PHASE_BAND = "FFFBF7EC"
INK = "FF212121"
MUTED = "FF6E6E6E"
WEEKLINE = "FFD9D2C2"
TODAY = "FFAD3815"

GRID_COL0 = 6  # coluna F = 1º dia da grade (A..E = identidade/colunas fixas)
MAX_WEEKS = 130  # teto defensivo (~2,5 anos); além disso, trunca COM aviso visível na planilha
COL_DIA = 2.6  # largura de cada coluna-dia
MES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
DIA_INI = ["S", "T", "Q", "Q", "S", "S", "D"]  # seg..dom (date.weekday(): seg=0)

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")

# Estilos reaproveitados (openpyxl deduplica, mas instanciar 1x por célula é caro em obra grande):
# uma instância por cor/borda, reusada em todas as células.
_FILLS = {
    c: PatternFill("solid", fgColor=c)
    for c in (GOLD, GOLD_DONE, LATE, LATE_DONE, DONE, WEEKEND, PHASE_BAND)
}
_BORD_SEMANA = Border(left=Side(style="thin", color=WEEKLINE))
_BORD_HOJE = Border(left=Side(style="medium", color=TODAY))
_BORD_SWATCH = Border(*(Side(style="thin", color=WEEKLINE),) * 4)  # amostra de cor da legenda


def _data(d):
    """Normaliza para date (get_tree devolve date; concluida_em pode vir datetime)."""
    return d.date() if isinstance(d, dt.datetime) else d


def _folhas(t: dict) -> list[dict]:
    subs = t.get("subitens") or []
    return subs if subs else [t]


def _prog_folha(lf: dict) -> float:
    """Avanço 0..1 da folha: medição (progresso_pct) se houver; senão binário do estado."""
    p = lf.get("progresso_pct")
    if p is not None:
        return max(0.0, min(1.0, float(p) / 100.0))
    return 1.0 if lf.get("estado") == "concluido" else 0.0


def _prog_tarefa(t: dict) -> float:
    fs = _folhas(t)
    return sum(_prog_folha(f) for f in fs) / len(fs) if fs else 0.0


def _tarefas_da_etapa(e: dict) -> list[dict]:
    """Tarefas da etapa na ordem da tela: subetapas ANTES das diretas (espelha tarefasDaEtapa)."""
    out: list[dict] = []
    for s in e.get("subetapas") or []:
        out += s.get("itens") or []
    out += e.get("itens") or []
    return out


def _prog_etapa(e: dict) -> float | None:
    """Avanço ponderado da etapa (espelha contagemEtapa): folhas das tarefas + cada subetapa-marco
    como 1 unidade (feita se concluída). None se não há nada a medir."""
    folhas = [lf for t in _tarefas_da_etapa(e) for lf in _folhas(t)]
    total = len(folhas)
    soma = sum(_prog_folha(lf) for lf in folhas)
    for s in e.get("subetapas") or []:
        if not (s.get("itens") or []):
            total += 1
            soma += 1.0 if s.get("concluida") else 0.0
    return soma / total if total else None


def _status(fim, prog, hoje) -> str:
    """concluido (>=100%) · atrasado (venceu e não fechou) · normal (previsto/andamento).
    Espelha statusDe (gantt.ts): sem nada a medir (prog None) NUNCA fica vermelho."""
    if prog is None:
        return "normal"
    if prog >= 1.0:
        return "concluido"
    if fim and fim < hoje:
        return "atrasado"
    return "normal"


def _cores(status: str) -> tuple[str, str | None]:
    """(cor da barra, cor da fração concluída). Concluído = barra inteira verde (sem fração)."""
    if status == "concluido":
        return DONE, None
    if status == "atrasado":
        return LATE, LATE_DONE
    return GOLD, GOLD_DONE


def _linhas(etapas: list[dict], hoje) -> tuple[list[dict], list]:
    """Achata a árvore em linhas desenháveis (etapa-resumo + tarefas com datas), espelhando
    montarGantt. Retorna (linhas, todas_as_datas)."""
    linhas: list[dict] = []
    datas: list = []
    for e in etapas:
        tarefas = [
            t
            for t in _tarefas_da_etapa(e)
            if t.get("data_inicio") and t.get("data_fim")
        ]
        if tarefas:
            ini = min(t["data_inicio"] for t in tarefas)
            fim = max(t["data_fim"] for t in tarefas)
        elif e.get("data_inicio") and e.get("data_fim"):
            ini, fim = e["data_inicio"], e["data_fim"]
        elif e.get("sem_itens") and e.get("concluida"):
            ini = fim = _data(e.get("concluida_em")) or hoje
        else:
            continue  # nada agendado/desenhável nesta etapa

        if not e.get("sem_itens"):
            pe = _prog_etapa(e)
            se = _status(fim, pe, hoje)
        elif e.get("concluida"):
            pe, se = 1.0, "concluido"
        elif fim < hoje:
            pe, se = None, "atrasado"
        else:
            pe, se = None, "normal"

        datas += [ini, fim]
        linhas.append({
            "kind": "etapa", "nome": e.get("nome") or "", "seq": e.get("seq_humano"),
            "inicio": ini, "fim": fim, "progresso": pe, "status": se,
        })
        for t in tarefas:
            pt = _prog_tarefa(t)
            ti, tf = t["data_inicio"], t["data_fim"]
            datas += [ti, tf]
            linhas.append({
                "kind": "tarefa", "nome": t.get("nome") or "", "seq": t.get("seq_humano"),
                "inicio": ti, "fim": tf, "progresso": pt, "status": _status(tf, pt, hoje),
            })
    return linhas, datas


def _save(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_cronograma_xlsx(
    obra: dict,
    etapas: list[dict],
    empresa: str | None,
    arquiteto: str | None,
    gerado_em: str,
    hoje: dt.date | None = None,
) -> bytes:
    """Monta o cronograma (.xlsx). `etapas` é a árvore de get_tree. `hoje` injetável p/ testes."""
    hoje = hoje or dt.date.today()
    linhas, datas = _linhas(etapas, hoje)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"
    for col, w in {"A": 42, "B": 6, "C": 11, "D": 11, "E": 6}.items():
        ws.column_dimensions[col].width = w

    # ---- banda de identidade (linhas 1-3, congeladas) ----
    _texto(ws, 1, 1, empresa or "", Font(bold=True, size=12, color=GOLD))
    ws.merge_cells("A1:E1")
    _texto(ws, 2, 1, obra.get("nome") or "Obra", Font(bold=True, size=16, color=INK))
    ws.merge_cells("A2:E2")
    seq = obra.get("seq_humano")
    arq = f"Arquiteto: {arquiteto}" if arquiteto else ""
    prefixo_obra = f"Obra #{seq}" if seq is not None else ""
    sub3 = " · ".join(p for p in (prefixo_obra, arq) if p)
    _texto(ws, 3, 1, sub3, Font(size=9, color=MUTED))
    ws.merge_cells("A3:E3")

    # títulos das colunas fixas (linha 5, ao lado dos cabeçalhos de semana)
    for i, tx in enumerate(["Tarefa / Fase", "%", "Início", "Término", "Dias"], start=1):
        _texto(ws, 5, i, tx, Font(bold=True, size=9, color=MUTED),
               _LEFT if i == 1 else _CENTER)
    ws.freeze_panes = "F7"
    ws.row_dimensions[2].height = 22

    if not datas:  # nenhuma data → planilha mínima (sem grade), sem quebrar
        _texto(ws, 7, 1, "Sem datas de cronograma para exibir.", Font(italic=True, color=MUTED))
        return _save(wb)

    window_min, window_max = min(datas), max(datas)
    grid_start = window_min - dt.timedelta(days=window_min.weekday())  # 2ª-feira da 1ª semana
    weeks_reais = (window_max - grid_start).days // 7 + 1
    truncado = weeks_reais > MAX_WEEKS  # obra além do teto: barras seriam cortadas
    weeks = min(weeks_reais, MAX_WEEKS)
    total_days = weeks * 7
    grid_end = grid_start + dt.timedelta(days=total_days - 1)

    dias_obra = (window_max - window_min).days + 1
    _texto(
        ws, 4, 1,
        f"Início {window_min:%d/%m/%Y} · Fim previsto {window_max:%d/%m/%Y} · {dias_obra} dias",
        Font(size=9, color=MUTED),
    )
    ws.merge_cells("A4:E4")
    _texto(ws, 1, GRID_COL0, f"Gerado em {gerado_em}", Font(size=8, color=MUTED))

    # ---- legenda das cores (linha 2; mesma semântica do Gantt da tela) ----
    leg_col = GRID_COL0
    for cor, rotulo in ((GOLD, "Previsto / em andamento"), (DONE, "Concluído"), (LATE, "Atrasado")):
        sw = ws.cell(row=2, column=leg_col)
        sw.fill = _FILLS[cor]
        sw.border = _BORD_SWATCH
        lab = ws.cell(row=2, column=leg_col + 1, value=rotulo)
        lab.font = Font(size=8, color=INK)
        lab.alignment = _LEFT
        ws.merge_cells(start_row=2, start_column=leg_col + 1, end_row=2, end_column=leg_col + 9)
        leg_col += 11
    nota = ws.cell(row=2, column=leg_col, value="Parte escura da barra = % já concluído")
    nota.font = Font(size=8, italic=True, color=MUTED)
    nota.alignment = _LEFT
    ws.merge_cells(start_row=2, start_column=leg_col, end_row=2, end_column=leg_col + 14)

    if truncado:  # corte NÃO-silencioso: avisa que a grade não cobre a obra inteira
        _texto(
            ws, 3, GRID_COL0,
            f"Cronograma além de {MAX_WEEKS} semanas — grade exibida só até {grid_end:%d/%m/%Y}",
            Font(size=9, bold=True, color=LATE),
        )
        ws.merge_cells(start_row=3, start_column=GRID_COL0, end_row=3, end_column=GRID_COL0 + 13)

    for c in range(total_days):
        ws.column_dimensions[get_column_letter(GRID_COL0 + c)].width = COL_DIA

    # ---- cabeçalhos da grade: meses (linha 4) · semanas (5) · inicial do dia (6) ----
    i = 0
    while i < total_days:
        d = grid_start + dt.timedelta(days=i)
        j = i
        while j < total_days:
            dj = grid_start + dt.timedelta(days=j)
            if dj.month != d.month or dj.year != d.year:
                break
            j += 1
        ws.merge_cells(
            start_row=4, start_column=GRID_COL0 + i, end_row=4, end_column=GRID_COL0 + j - 1
        )
        mc = ws.cell(row=4, column=GRID_COL0 + i, value=f"{MES[d.month - 1]}/{d.year % 100:02d}")
        mc.font = Font(bold=True, size=9, color=INK)
        mc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        i = j
    for w in range(weeks):
        c0 = GRID_COL0 + w * 7
        ws.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c0 + 6)
        wc = ws.cell(row=5, column=c0, value=f"Semana {w + 1}")
        wc.font = Font(size=8, color=MUTED)
        wc.alignment = _CENTER
    for c in range(total_days):
        d = grid_start + dt.timedelta(days=c)
        cell = ws.cell(row=6, column=GRID_COL0 + c, value=DIA_INI[d.weekday()])
        cell.font = Font(size=7, color=MUTED)
        cell.alignment = _CENTER
        if d.weekday() >= 5:
            cell.fill = _FILLS[WEEKEND]

    # Linha de HOJE só dentro da janela EFETIVA da obra (não no padding da grade) — igual à tela.
    today_col = (
        GRID_COL0 + (hoje - grid_start).days if window_min <= hoje <= window_max else None
    )

    # ---- corpo: uma linha por etapa-resumo / tarefa ----
    r = 7
    for ln in linhas:
        is_etapa = ln["kind"] == "etapa"
        rotulo = (f"#{ln['seq']} " if ln["seq"] is not None else "") + ln["nome"]
        nc = ws.cell(row=r, column=1, value=rotulo)
        nc.font = Font(bold=is_etapa, size=10 if is_etapa else 9, color=INK)
        nc.alignment = Alignment(vertical="center", indent=0 if is_etapa else 1)
        if ln["progresso"] is not None:
            pc = ws.cell(row=r, column=2, value=round(ln["progresso"], 4))
            pc.number_format = "0%"
            pc.font = Font(size=9, color=MUTED)
            pc.alignment = _CENTER
        for col, val in ((3, ln["inicio"]), (4, ln["fim"])):
            dc = ws.cell(row=r, column=col, value=val)
            dc.number_format = "DD/MM/YYYY"
            dc.font = Font(size=9, color=MUTED)
            dc.alignment = _CENTER
        diasc = ws.cell(row=r, column=5, value=(ln["fim"] - ln["inicio"]).days + 1)
        diasc.font = Font(size=9, color=MUTED)
        diasc.alignment = _CENTER
        if is_etapa:
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = _FILLS[PHASE_BAND]

        base, frac = _cores(ln["status"])
        span = (ln["fim"] - ln["inicio"]).days + 1
        feitos = span if ln["status"] == "concluido" else round((ln["progresso"] or 0.0) * span)
        off0 = (ln["inicio"] - grid_start).days
        for k in range(span):
            off = off0 + k
            if not (0 <= off < total_days):
                continue
            cor = frac if (frac and k < feitos) else base
            ws.cell(row=r, column=GRID_COL0 + off).fill = _FILLS[cor]
        ws.row_dimensions[r].height = 15
        r += 1

    # ---- separadores de semana + linha de hoje (por cima dos preenchimentos) ----
    for rr in range(4, r):
        for w in range(weeks):
            ws.cell(row=rr, column=GRID_COL0 + w * 7).border = _BORD_SEMANA
        if today_col is not None:
            ws.cell(row=rr, column=today_col).border = _BORD_HOJE

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:6"
    ws.print_title_cols = "A:E"
    return _save(wb)


def _texto(ws, row: int, col: int, value: str, font: Font, align: Alignment = _LEFT) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = align
