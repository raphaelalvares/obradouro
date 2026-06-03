"""Testes da Fase 3 (checklist) sem banco: normalização (chave de dedupe) e parser do template."""

import io

import openpyxl
import pytest
from fastapi import HTTPException

from app.services.checklist_import import (
    TEMPLATE_HEADER,
    _num,
    norm_nome,
    parse_template,
    parse_xlsx,
)


def _xlsx(rows: list[tuple]) -> bytes:
    """Monta um .xlsx em memória: 1ª linha = cabeçalho do template, demais = dados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(TEMPLATE_HEADER))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------- norm_nome (contrato congelado de dedupe) ----------------
def test_norm_nome_dobra_acento_e_caixa():
    # acento, caixa e espaços: tudo colapsa na MESMA chave (dedupe estável no import/create)
    assert norm_nome("Fundação") == norm_nome("fundacao") == norm_nome("  FUNDACAO ")
    assert norm_nome("Acabamento   Final") == "acabamento final"


def test_norm_nome_forma_unicode():
    composto = "Fundação"  # ç,ã compostos (NFC)
    decomposto = "Fundação"  # c+cedilha, a+til (NFD)
    assert norm_nome(composto) == norm_nome(decomposto)


def test_norm_nome_vazio():
    assert norm_nome("") == ""
    assert norm_nome("   ") == ""


# ---------------- parse_template ----------------
def test_parse_template_arvore_e_forward_fill():
    raw = _xlsx(
        [
            ("Fundação", "Sapatas", 1, 1),
            (None, "Vigas baldrame", None, 2),  # etapa em branco → herda 'Fundação'
            ("Alvenaria", "Blocos", 2, 1),
        ]
    )
    payload = parse_template(raw)
    assert [e["nome"] for e in payload] == ["Fundação", "Alvenaria"]
    fund = payload[0]
    assert fund["nome_norm"] == "fundacao"
    assert [i["nome"] for i in fund["itens"]] == ["Sapatas", "Vigas baldrame"]
    assert payload[1]["itens"][0]["nome"] == "Blocos"


def test_parse_template_dedup_no_arquivo():
    raw = _xlsx(
        [
            ("Pintura", "Massa corrida", 1, 1),
            ("Pintura", "massa  corrida", 1, 2),  # mesma etapa+item normalizados → 1ª vence
        ]
    )
    payload = parse_template(raw)
    assert len(payload) == 1
    assert len(payload[0]["itens"]) == 1


def test_parse_template_cabecalho_invalido_422():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["fase", "tarefa", "x", "y"])  # cabeçalho fora do padrão
    ws.append(["Fundação", "Sapatas", 1, 1])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(HTTPException) as exc:
        parse_template(buf.getvalue())
    assert exc.value.status_code == 422


def test_parse_template_item_sem_etapa_nao_some_silenciosamente():
    raw = _xlsx([(None, "Item órfão", None, 1)])  # item sem etapa anterior → erro, não drop
    with pytest.raises(HTTPException) as exc:
        parse_template(raw)
    assert exc.value.status_code == 422
    assert "sem etapa" in exc.value.detail


def test_parse_template_nao_e_xlsx_422():
    with pytest.raises(HTTPException) as exc:
        parse_template(b"isto nao e um arquivo xlsx")
    assert exc.value.status_code == 422


# ---------------- _num (valores do orçamento) ----------------
def test_num_formatos():
    assert _num(600) == 600.0
    assert _num(4233.44) == 4233.44
    assert _num("600") == 600.0
    assert _num("R$9.000,00") == 9000.0  # BR: ponto=milhar, vírgula=decimal
    assert _num("1.234,5") == 1234.5
    assert _num("4233.44") == 4233.44  # ponto-decimal puro
    assert _num("X") is None and _num("") is None and _num(None) is None
    assert _num("lixo") is None  # não explode


# ---------------- parse_xlsx: planilha de ORÇAMENTO real ----------------
def _orcamento_xlsx(linhas: list[tuple]) -> bytes:
    """Monta um .xlsx no formato do orçamento real: linhas de topo + cabeçalho ITEM/DESCRIÇÃO/..."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ESTUDO DE VIABILIDADE FINANCEIRA"])  # título (linha de topo, ignorada)
    ws.append(["Cliente Fulano"])  # idem
    ws.append(
        ["ITEM", "DESCRIÇÃO DOS SERVIÇOS", "UNIDADE", "QUANTIDADE", "M.O", "MAT.", "TOTAL"]
    )
    for r in linhas:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_orcamento_etapas_itens_valores():
    raw = _orcamento_xlsx(
        [
            ("01", "SERVIÇOS PRELIMINARES"),
            ("01.01", "Projetos complementares", "unidade", 0, "X", "X", 0),
            ("01.02", "ART", "Verba", 1, "X", "600", 600),
            ("SUBTOTAL DO ITEM 01", None, None, None, None, None, 600),
            ("02", "REMOÇÕES E DEMOLIÇÃO"),
            (None, "GERAL"),  # sub-rótulo (só B) → ignorado
            ("02.01", "Remoção de piso", "m²", 75, 5000, "X", 5000),
            ("02.01", "remoção  de piso", "m²", 75, 5000, "X", 5000),  # dup norm → 1ª vence
            ("TOTAL", None, None, None, None, None, 87368),
            ("1. Nota qualquer do rodapé que deve ser ignorada",),  # nota → ignorada
        ]
    )
    payload = parse_xlsx(raw)
    assert [e["nome"] for e in payload] == ["SERVIÇOS PRELIMINARES", "REMOÇÕES E DEMOLIÇÃO"]

    e1 = payload[0]
    assert [i["nome"] for i in e1["itens"]] == ["Projetos complementares", "ART"]
    art = e1["itens"][1]
    assert art["unidade"] == "Verba"
    assert art["quantidade"] == 1.0
    assert art["custo_mao_obra"] is None  # "X"
    assert art["custo_material"] == 600.0
    assert art["custo_total"] == 600.0
    assert art["ambiente"] is None  # orçamento não tem cômodo

    e2 = payload[1]
    assert len(e2["itens"]) == 1  # dedup por nome_norm
    piso = e2["itens"][0]
    assert piso["nome"] == "Remoção de piso"
    assert piso["quantidade"] == 75.0
    assert piso["custo_mao_obra"] == 5000.0
    assert piso["custo_material"] is None  # "X"


def test_parse_xlsx_roteia_template_tambem():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(TEMPLATE_HEADER))
    ws.append(["Fundação", "Sapatas", 1, 1])
    buf = io.BytesIO()
    wb.save(buf)
    payload = parse_xlsx(buf.getvalue())
    assert payload[0]["nome"] == "Fundação"
    assert payload[0]["itens"][0]["nome"] == "Sapatas"


def test_parse_xlsx_formato_desconhecido_422():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["qualquer", "coisa", "aqui"])
    ws.append(["sem", "padrão", "nenhum"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(HTTPException) as exc:
        parse_xlsx(buf.getvalue())
    assert exc.value.status_code == 422
