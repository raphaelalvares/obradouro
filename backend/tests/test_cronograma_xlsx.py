"""Testes do renderizador de cronograma Excel (puro, sem DB). Gera os bytes, reabre com openpyxl e
confere cabeçalho, faixa de semanas/meses e as cores das barras (situação + fração de avanço)."""

import datetime as dt
from io import BytesIO

from openpyxl import load_workbook

from app.services.xlsx_render import (
    DONE,
    GOLD,
    GOLD_DONE,
    LATE,
    render_cronograma_xlsx,
)


def _etapa(nome, seq, itens, **kw):
    base = {
        "nome": nome, "seq_humano": seq, "sem_itens": False, "concluida": False,
        "concluida_em": None, "data_inicio": None, "data_fim": None,
        "subetapas": [], "itens": itens,
    }
    base.update(kw)
    return base


def _tarefa(nome, seq, ini, fim, estado="pendente", prog=None, subitens=None):
    return {
        "nome": nome, "seq_humano": seq, "data_inicio": ini, "data_fim": fim,
        "estado": estado, "progresso_pct": prog, "subitens": subitens or [],
    }


def _gerar(etapas, hoje, obra=None, empresa="Emp", arquiteto="Arq"):
    data = render_cronograma_xlsx(
        obra or {"nome": "Obra X", "seq_humano": 7}, etapas, empresa, arquiteto, "01/01 00:00",
        hoje=hoje,
    )
    return load_workbook(BytesIO(data)).active


def _gs(d: dt.date) -> dt.date:
    """2ª-feira da semana de d (mesmo cálculo do renderer)."""
    return d - dt.timedelta(days=d.weekday())


def _col(gs: dt.date, d: dt.date) -> int:
    return 6 + (d - gs).days


def test_cabecalho_empresa_obra_arquiteto():
    ws = _gerar([_etapa("E1", 1, [_tarefa("T1", 2, dt.date(2026, 6, 8), dt.date(2026, 6, 9))])],
                dt.date(2026, 6, 1))
    assert ws["A1"].value == "Emp"
    assert ws["A2"].value == "Obra X"
    assert "Arq" in ws["A3"].value
    assert "Obra #7" in ws["A3"].value


def test_semanas_e_mes():
    ini, fim = dt.date(2026, 6, 8), dt.date(2026, 6, 20)  # 2ª-feira → grid começa nele; 13 dias
    ws = _gerar([_etapa("E", 1, [_tarefa("T", 2, ini, fim)])], dt.date(2026, 6, 1))
    assert ws.cell(row=5, column=6).value == "Semana 1"
    assert ws.cell(row=5, column=13).value == "Semana 2"  # 7 colunas adiante
    assert ws.cell(row=4, column=6).value == "JUN/26"


def test_barra_prevista_amarela():
    ini, fim = dt.date(2026, 6, 8), dt.date(2026, 6, 10)  # 3 dias, no futuro
    ws = _gerar([_etapa("E", 1, [_tarefa("T", 2, ini, fim)])], dt.date(2026, 6, 1))
    gs = _gs(ini)
    for d in (ini, ini + dt.timedelta(days=1), fim):
        assert ws.cell(row=8, column=_col(gs, d)).fill.fgColor.rgb == GOLD
    # fora da barra → sem preenchimento
    assert ws.cell(row=8, column=_col(gs, fim + dt.timedelta(days=1))).fill.patternType is None


def test_barra_concluida_verde():
    ini, fim = dt.date(2026, 6, 8), dt.date(2026, 6, 10)
    ws = _gerar(
        [_etapa("E", 1, [_tarefa("T", 2, ini, fim, estado="concluido", prog=100)])],
        dt.date(2026, 6, 20),
    )
    gs = _gs(ini)
    assert ws.cell(row=8, column=_col(gs, ini)).fill.fgColor.rgb == DONE


def test_barra_atrasada_vermelha():
    ini, fim = dt.date(2026, 6, 8), dt.date(2026, 6, 10)
    ws = _gerar([_etapa("E", 1, [_tarefa("T", 2, ini, fim)])], dt.date(2026, 7, 1))  # já venceu
    gs = _gs(ini)
    assert ws.cell(row=8, column=_col(gs, ini)).fill.fgColor.rgb == LATE


def test_fracao_de_avanco():
    ini, fim = dt.date(2026, 6, 8), dt.date(2026, 6, 11)  # 4 dias
    ws = _gerar(
        [_etapa("E", 1, [_tarefa("T", 2, ini, fim, estado="em_andamento", prog=50)])],
        dt.date(2026, 6, 1),
    )
    gs = _gs(ini)
    # 50% de 4 dias = 2 células concluídas (tom escuro), 2 previstas (âmbar)
    assert ws.cell(row=8, column=_col(gs, ini)).fill.fgColor.rgb == GOLD_DONE
    assert ws.cell(row=8, column=_col(gs, fim)).fill.fgColor.rgb == GOLD


def test_sem_datas_planilha_minima():
    ws = _gerar(
        [_etapa("Marco", 1, [], sem_itens=True)],
        dt.date(2026, 6, 1),
    )
    assert "Sem datas" in ws["A7"].value


def test_datas_invertidas_nao_quebra():
    # término antes do início (dado ruim do usuário) → não desenha barra, mas NÃO estoura.
    ini, fim = dt.date(2026, 6, 10), dt.date(2026, 6, 5)
    ws = _gerar([_etapa("E", 1, [_tarefa("T", 2, ini, fim)])], dt.date(2026, 6, 1))
    assert ws["A2"].value == "Obra X"  # gerou a planilha sem exceção
